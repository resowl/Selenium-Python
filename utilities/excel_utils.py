import openpyxl


def get_row_count(path, sheet_name):
    wb = openpyxl.load_workbook(path)
    sh = wb[sheet_name]
    return sh.max_row


def get_column_count(path, sheet_name):
    wb = openpyxl.load_workbook(path)
    sh = wb[sheet_name]
    return sh.max_column


def read_data(path, sheet_name, row_num, column_num):
    wb = openpyxl.load_workbook(path)
    sh = wb[sheet_name]
    val = sh.cell(row=row_num, column=column_num).value
    if val is None:
        val = ""
    return val


def write_data(path, sheet_name, row_num, column_num, data):
    wb = openpyxl.load_workbook(path)
    sh = wb[sheet_name]
    sh.cell(row=row_num, column=column_num).value = data
    wb.save(path)